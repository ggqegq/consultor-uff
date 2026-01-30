# ==============================================
# CONSULTOR DE QUADRO DE HORÁRIOS UFF
# Versão SEM Selenium - usa requests + BeautifulSoup
# Compatível com Streamlit Cloud
# ==============================================

import streamlit as st
import pandas as pd
import requests
from bs4 import BeautifulSoup
import re
import io
import time

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ===== CONFIGURAÇÃO DA PÁGINA =====
st.set_page_config(
    page_title="Consultor UFF - Quadro de Horários",
    page_icon="📊",
    layout="wide"
)

# ===== ESTILOS CSS =====
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1e3a5f;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #666;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        padding: 1rem;
        background-color: #d4edda;
        border-radius: 0.5rem;
        border-left: 4px solid #28a745;
    }
    .info-box {
        padding: 1rem;
        background-color: #e7f3ff;
        border-radius: 0.5rem;
        border-left: 4px solid #0066cc;
    }
</style>
""", unsafe_allow_html=True)

# ===== FUNÇÕES AUXILIARES =====
def calcular_periodos_retroativos(periodo_base, qtd=3):
    """Gera lista de períodos a partir de uma base."""
    periodo_base = periodo_base.replace('.', '')
    ano = int(periodo_base[:4])
    semestre = int(periodo_base[4])
    lista_periodos = []
    for _ in range(qtd):
        lista_periodos.append(f"{ano}{semestre}")
        if semestre == 1:
            semestre = 2
            ano -= 1
        else:
            semestre = 1
    return lista_periodos

# ===== CLASSE PRINCIPAL (SEM SELENIUM) =====
class ConsultorQuadroHorariosUFF:
    def __init__(self, periodos, curso_filtro=None, departamentos_filtro=None):
        self.periodos = periodos
        self.curso_filtro = curso_filtro
        self.departamentos_filtro = departamentos_filtro if departamentos_filtro else []
        self.links_processados = set()
        
        # Sessão HTTP com headers de navegador
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        
        self.ids_cursos = {
            'Química': '28',
            'Química Industrial': '29'
        }

    def construir_url_busca(self, id_curso, departamento=None, periodo='20252', pagina=1):
        base_url = "https://app.uff.br/graduacao/quadrodehorarios/"
        params = [
            "utf8=%E2%9C%93",
            f"q%5Banosemestre_eq%5D={periodo}",
            "q%5Bdisciplina_cod_departamento_eq%5D=",
            f"q%5Bvagas_turma_curso_idcurso_eq%5D={id_curso}",
        ]
        if departamento and departamento.strip():
            codigo_busca = f"{departamento.strip().upper()}00"
            params.insert(0, f"q%5Bdisciplina_nome_or_disciplina_codigo_cont%5D={codigo_busca}")
        else:
            params.insert(0, "q%5Bdisciplina_nome_or_disciplina_codigo_cont%5D=")
        
        if pagina > 1:
            params.append(f"page={pagina}")
            
        return base_url + "?" + "&".join(params)

    def extrair_links_turmas_da_pagina(self, html):
        """Extrai links de turmas do HTML da página."""
        links = set()
        try:
            soup = BeautifulSoup(html, 'html.parser')
            for link in soup.find_all('a', href=True):
                if '/turmas/' in link['href']:
                    full_url = f"https://app.uff.br{link['href']}" if not link['href'].startswith('http') else link['href']
                    links.add(full_url.split('?')[0])
        except Exception as e:
            st.warning(f"Erro ao extrair links: {e}")
        return list(links)

    def tem_proxima_pagina(self, html):
        """Verifica se existe próxima página na paginação."""
        try:
            soup = BeautifulSoup(html, 'html.parser')
            paginacao = soup.find('nav', class_='pagination') or soup.find('ul', class_='pagination')
            if paginacao:
                next_link = paginacao.find('a', rel='next') or paginacao.find('a', string=re.compile('›|Próximo|Next'))
                return next_link is not None
        except:
            pass
        return False

    def navegar_todas_paginas(self, id_curso, departamento, periodo):
        """Navega por todas as páginas e coleta links."""
        todos_links = set()
        pagina = 1
        max_paginas = 50  # Limite de segurança
        
        while pagina <= max_paginas:
            url = self.construir_url_busca(id_curso, departamento, periodo, pagina)
            
            try:
                response = self.session.get(url, timeout=30)
                response.raise_for_status()
                html = response.text
                
                # Verificar se a página carregou corretamente
                if 'quadrodehorarios' not in html.lower() and len(html) < 1000:
                    st.warning(f"Página pode estar incompleta (página {pagina})")
                    break
                
                links = self.extrair_links_turmas_da_pagina(html)
                
                if not links:
                    break
                    
                todos_links.update(links)
                
                if not self.tem_proxima_pagina(html):
                    break
                    
                pagina += 1
                time.sleep(0.3)  # Pequena pausa para não sobrecarregar o servidor
                
            except requests.exceptions.RequestException as e:
                st.error(f"Erro de conexão na página {pagina}: {e}")
                break
                
        return list(todos_links)

    def extrair_dados_turma_por_curso(self, url_turma, periodo, curso_alvo):
        """Extrai dados de uma turma específica para um curso específico."""
        try:
            response = self.session.get(url_turma, timeout=30)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extrair título
            h1 = soup.find('h1')
            if not h1:
                return None
                
            titulo = h1.get_text(strip=True)
            match = re.search(r'Turma\s+(\S+)\s+de\s+(\S+)\s+-\s+(.+)', titulo)
            if not match:
                return None
            
            turma, codigo, nome = match.group(1), match.group(2), match.group(3)
            depto = codigo[:3]
            
            # Extrair horários
            horario_str = "Não informado"
            try:
                h5_horario = soup.find('h5', string=re.compile('Horários'))
                if h5_horario:
                    tabela = h5_horario.find_next('table')
                    if tabela:
                        trs = tabela.find_all('tr')
                        if len(trs) > 1:
                            cols = trs[1].find_all('td')
                            dias = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb']
                            horarios = [f"{dias[i]}: {c.text.strip()}" for i, c in enumerate(cols) if c.text.strip() and i < 6]
                            if horarios:
                                horario_str = " | ".join(horarios)
            except Exception:
                pass
            
            # Extrair vagas para o curso específico
            vagas_info = None
            try:
                h5_vagas = soup.find('h5', string=re.compile('Vagas Alocadas'))
                if h5_vagas:
                    tabela = h5_vagas.find_next('table')
                    if tabela:
                        trs = tabela.find_all('tr')[2:]  # Pular cabeçalhos
                        for row in trs:
                            cols = row.find_all('td')
                            if len(cols) >= 5:
                                curso_nome = cols[0].text.strip()
                                match_curso = False
                                
                                # Verificação mais precisa do curso
                                if curso_alvo == 'Química':
                                    if '028' in curso_nome or ('Química' in curso_nome and 'Industrial' not in curso_nome):
                                        match_curso = True
                                elif curso_alvo == 'Química Industrial':
                                    if '029' in curso_nome or 'Industrial' in curso_nome:
                                        match_curso = True
                                    
                                if match_curso:
                                    vagas_info = {
                                        'vagas_reg': int(cols[1].text) if cols[1].text.strip().isdigit() else 0,
                                        'vagas_vest': int(cols[2].text) if cols[2].text.strip().isdigit() else 0,
                                        'inscritos_reg': int(cols[3].text) if cols[3].text.strip().isdigit() else 0,
                                        'inscritos_vest': int(cols[4].text) if cols[4].text.strip().isdigit() else 0,
                                    }
                                    break
            except Exception:
                pass
            
            if not vagas_info:
                return None
            
            return {
                'periodo': periodo,
                'curso': curso_alvo,  # Adicionando o curso nos dados
                'depto': depto,
                'codigo': codigo,
                'disciplina': nome,
                'turma': turma,
                'horario': horario_str,
                **vagas_info
            }
        except Exception as e:
            return None

    def executar_consulta(self, progress_bar, status_text):
        """Executa a consulta completa."""
        dados_brutos = []
        cursos_para_buscar = [self.curso_filtro] if self.curso_filtro else list(self.ids_cursos.keys())
        total_steps = len(self.periodos) * len(cursos_para_buscar)
        step = 0
        
        for periodo in self.periodos:
            for curso in cursos_para_buscar:
                step += 1
                progress = step / total_steps * 0.5  # Primeira metade: coleta de links
                progress_bar.progress(progress)
                status_text.text(f"Buscando {curso} em {periodo[:4]}.{periodo[4]}...")
                
                # Criar um set separado para controlar links já processados NESTE curso/período
                links_processados_neste_ciclo = set()
                
                id_curso = self.ids_cursos.get(curso, '28')
                deptos = self.departamentos_filtro if self.departamentos_filtro else [None]
                
                todos_links = []
                for depto in deptos:
                    links = self.navegar_todas_paginas(id_curso, depto, periodo)
                    todos_links.extend(links)
                
                # Remover duplicatas
                todos_links = list(set(todos_links))
                
                # Processar cada turma
                total_links = len(todos_links)
                for idx, link in enumerate(todos_links):
                    progress = 0.5 + (step / total_steps * 0.5) * (idx / max(total_links, 1))
                    progress_bar.progress(min(progress, 0.99))
                    status_text.text(f"Processando turma {idx+1}/{total_links} de {curso} ({periodo[:4]}.{periodo[4]})...")
                    
                    # Evitar processar o mesmo link duas vezes no mesmo ciclo curso/período
                    if link in links_processados_neste_ciclo:
                        continue
                    links_processados_neste_ciclo.add(link)
                    
                    dado = self.extrair_dados_turma_por_curso(link, periodo, curso)
                    if dado:
                        dados_brutos.append(dado)
                    
                    time.sleep(0.2)  # Pausa para não sobrecarregar
        
        return dados_brutos

    def gerar_excel_comparativo(self, dados):
        """Gera planilha Excel comparativa."""
        if not dados:
            return None
        
        df = pd.DataFrame(dados)
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo de Períodos"
        
        blue_fill = PatternFill(start_color="337AB7", end_color="337AB7", fill_type="solid")
        beige_fill = PatternFill(start_color="FDFDF0", end_color="FDFDF0", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        periodos_ordenados = sorted(list(df['periodo'].unique()), reverse=True)
        sub_cols = [
            ('horario', 'Horário'),
            ('vagas_reg', 'Vagas Reg'),
            ('inscritos_reg', 'Insc Reg'),
            ('vagas_vest', 'Vagas Vest'),
            ('inscritos_vest', 'Insc Vest')
        ]
        
        # Cabeçalhos fixos
        headers_row1 = ["Curso", "Depto", "Código", "Disciplina", "Turma"]
        for col_idx, text in enumerate(headers_row1, 1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        
        # Cabeçalhos de períodos
        current_col = 6
        for per in periodos_ordenados:
            per_fmt = f"{per[:4]}.{per[4]}"
            cell = ws.cell(row=1, column=current_col, value=per_fmt)
            ws.merge_cells(start_row=1, start_column=current_col, end_row=1, end_column=current_col + len(sub_cols) - 1)
            cell.fill = blue_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            
            for _, title in sub_cols:
                sub_cell = ws.cell(row=2, column=current_col, value=title)
                sub_cell.fill = blue_fill
                sub_cell.font = Font(color="FFFFFF", bold=False, size=9)
                sub_cell.alignment = center
                sub_cell.border = border
                current_col += 1
        
        # Dados - agrupando por curso também
        grouped = df.groupby(['curso', 'depto', 'codigo', 'disciplina', 'turma'])
        row_num = 3
        
        for name, group in grouped:
            for i, val in enumerate(name, 1):
                cell = ws.cell(row=row_num, column=i, value=val)
                cell.border = border
                cell.alignment = center if i != 4 else Alignment(horizontal='left', vertical='center')
            
            col_idx = 6
            for per in periodos_ordenados:
                dados_periodo = group[group['periodo'] == per]
                if not dados_periodo.empty:
                    dado = dados_periodo.iloc[0]
                    vals = [dado[k] for k, _ in sub_cols]
                else:
                    vals = ['-', '-', '-', '-', '-']
                
                for val in vals:
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    cell.border = border
                    cell.alignment = center
                    cell.fill = beige_fill
                    col_idx += 1
            row_num += 1
        
        # Ajustar larguras
        ws.column_dimensions['A'].width = 18  # Curso
        ws.column_dimensions['B'].width = 8   # Depto
        ws.column_dimensions['C'].width = 12  # Código
        ws.column_dimensions['D'].width = 35  # Disciplina
        ws.column_dimensions['E'].width = 8   # Turma
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


# ===== INTERFACE PRINCIPAL =====
st.markdown('<p class="main-header">Consultor de Quadro de Horários UFF</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Gere planilhas comparativas de vagas e horários dos cursos de Química</p>', unsafe_allow_html=True)

# Aviso sobre método
st.info("Para consultar um departamento específico digite  a sigla (Ex. GQA, GQI, GQO etc)!")

# Formulário
with st.form("consulta_form"):
    col1, col2 = st.columns(2)
    
    with col1:
        periodo_ref = st.text_input(
            "Período de Referência",
            value="2026.1",
            help="Formato: AAAA.S (ex: 2026.1)"
        )
        
        qtd_periodos = st.slider(
            "Quantidade de Períodos",
            min_value=1,
            max_value=6,
            value=3,
            help="Quantos períodos anteriores incluir na comparação"
        )
    
    with col2:
        curso = st.selectbox(
            "Curso",
            options=["Todos", "Química", "Química Industrial"],
            help="Selecione o curso ou deixe 'Todos'"
        )
        
        deptos = st.text_input(
            "Departamentos (opcional)",
            placeholder="Ex: GQI, GQO, TEP",
            help="Separe por vírgula. Deixe vazio para todos."
        )
    
    submitted = st.form_submit_button("Gerar Planilha", use_container_width=True)

# Processamento
if submitted:
    # Validação
    try:
        periodo_clean = periodo_ref.replace('.', '')
        if len(periodo_clean) != 5 or not periodo_clean.isdigit():
            raise ValueError("Formato inválido")
    except:
        st.error("Formato de período inválido. Use AAAA.S (ex: 2026.1)")
        st.stop()
    
    # Configurar parâmetros
    periodos = calcular_periodos_retroativos(periodo_ref, qtd_periodos)
    curso_filtro = curso if curso != "Todos" else None
    deptos_filtro = [d.strip().upper() for d in deptos.split(',')] if deptos.strip() else None
    
    # Mostrar configuração
    st.markdown("---")
    st.write(f"**Períodos a consultar:** {', '.join([f'{p[:4]}.{p[4]}' for p in periodos])}")
    st.write(f"**Curso:** {curso}")
    st.write(f"**Departamentos:** {deptos if deptos else 'Todos'}")
    
    # Barra de progresso
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        with st.spinner("Iniciando consulta..."):
            consultor = ConsultorQuadroHorariosUFF(periodos, curso_filtro, deptos_filtro)
            dados = consultor.executar_consulta(progress_bar, status_text)
        
        if dados:
            status_text.text("Gerando planilha Excel...")
            excel_buffer = consultor.gerar_excel_comparativo(dados)
            
            if excel_buffer:
                progress_bar.progress(1.0)
                status_text.text("Concluído!")
                
                st.success(f"Planilha gerada com sucesso! {len(dados)} registros encontrados.")
                
                nome_arquivo = f"Comparativo_{periodo_ref.replace('.','')}_e_anteriores.xlsx"
                
                st.download_button(
                    label="Download da Planilha Excel",
                    data=excel_buffer,
                    file_name=nome_arquivo,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.warning("Nenhum dado encontrado para os filtros selecionados. Isso pode significar que o site requer JavaScript para carregar os dados.")
            st.info("Se isso persistir, a alternativa é usar o Google Colab com Widgets, que suporta Selenium.")
            
    except Exception as e:
        st.error(f"Erro durante a consulta: {e}")
        st.info("Se o erro persistir, tente a versão para Google Colab que usa Selenium.")

# Rodapé
st.markdown("---")
st.markdown("""
<div style="text-align: center; color: #888; font-size: 0.9rem;">
    Desenvolvido por Tadeu L. Araujo<br>
    Versão 1.0 - compatível com Streamlit Cloud
</div>
""", unsafe_allow_html=True)
